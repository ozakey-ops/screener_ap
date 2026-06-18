"""
DART 재무제표 수집기 v1.0
────────────────────────────────────────────────────────────────
KRX 전체 종목의 최근 15년 사업보고서 재무데이터를 DART Open API로 수집

수집 항목 (이미지 기준):
  당기순이익 / 영업이익 / 매출액
  총자산 / 자기자본 / 총부채
  유동자산 / 유동부채
  이자비용 / 배당금 / 감가상각비 / 현금및현금성자산
  시가총액 (KRX 데이터)

API:
  fnlttMultiAcnt  → 핵심 재무항목 (다중 기업)
  fnlttCmpnyIndx  → 재무비율 지표 (다중 기업)
  corpCode.xml    → 기업코드 매핑

실행 순서:
  1. python krx_stock_collector.py   (KRX 종목 수집)
  2. python dart_financial_collector.py  (재무데이터 수집)

실행 방법:
  pip install requests pandas openpyxl
  python dart_financial_collector.py
────────────────────────────────────────────────────────────────
"""

import os
import sys
import glob
import time
import zipfile
import io
import json
import requests
import pandas as pd
from datetime import datetime
from xml.etree import ElementTree as ET
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
#  설정
# ═══════════════════════════════════════════════════════════════

DART_API_KEY   = "901de77da059b85e095a99ab9f2baf3264f7281f"
DART_BASE      = "https://opendart.fss.or.kr/api"
REPRT_CODE     = "11011"       # 사업보고서 (연간)
BATCH_SIZE     = 100           # API 1회 요청당 기업 수
REQUEST_DELAY  = 0.5           # 요청 간격 (초) - API 부하 방지
CHECKPOINT_DIR = "dart_checkpoint"   # 중간저장 폴더

CURRENT_YEAR   = datetime.today().year
TARGET_YEARS   = list(range(CURRENT_YEAR - 15, CURRENT_YEAR))  # 최근 15년

OUTPUT_FILE    = f"DART_재무데이터_{datetime.today().strftime('%Y%m%d')}.xlsx"

# 재무비율 지표 분류코드
IDX_CL_CODES = {
    "M210000": "수익성지표",
    "M220000": "안정성지표",
    "M230000": "성장성지표",
    "M240000": "활동성지표",
}

# fnlttMultiAcnt 계정명 → 표준 컬럼명
ACNT_MAP = {
    "매출액":               "매출액",
    "수익(매출액)":          "매출액",
    "영업수익":              "매출액",
    "영업이익":              "영업이익",
    "영업이익(손실)":        "영업이익",
    "법인세차감전계속사업이익": "세전이익",
    "법인세비용차감전순이익":  "세전이익",
    "당기순이익":            "당기순이익",
    "당기순이익(손실)":      "당기순이익",
    "자산총계":              "총자산",
    "부채총계":              "총부채",
    "자본총계":              "자기자본",
}

# 주요 재무항목 표시 순서
FIN_COLS_ORDER = [
    "매출액", "영업이익", "세전이익", "당기순이익",
    "총자산", "총부채", "자기자본",
    "유동자산", "유동부채",
    "이자비용", "배당금", "감가상각비", "현금및현금성자산",
    "시가총액",
]


# ═══════════════════════════════════════════════════════════════
#  유틸리티
# ═══════════════════════════════════════════════════════════════

def log(msg, indent=0):
    print("  " * indent + msg)


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def checkpoint_path(year, suffix="acnt"):
    return os.path.join(CHECKPOINT_DIR, f"{year}_{suffix}.csv")


def load_checkpoint(year, suffix="acnt"):
    p = checkpoint_path(year, suffix)
    if os.path.exists(p):
        return pd.read_csv(p, dtype=str)
    return None


def save_checkpoint(df, year, suffix="acnt"):
    ensure_dir(CHECKPOINT_DIR)
    df.to_csv(checkpoint_path(year, suffix), index=False, encoding="utf-8-sig")


# ═══════════════════════════════════════════════════════════════
#  STEP 1: KRX 종목 로드
# ═══════════════════════════════════════════════════════════════

def load_krx_stocks() -> pd.DataFrame:
    """
    krx_stock_collector.py가 생성한 Excel에서 종목코드·종목명 로드.
    파일이 없으면 data-dbg.krx.co.kr API로 직접 수집.
    """
    # 최신 KRX 파일 탐색
    files = sorted(glob.glob("KRX_종목목록_*.xlsx"), reverse=True)
    if files:
        log(f"KRX 파일 로드: {files[0]}")
        df = pd.read_excel(files[0], sheet_name="전체", dtype=str)
        df = df[["종목코드", "종목명", "시장구분"]].dropna(subset=["종목코드"])
        df["종목코드"] = df["종목코드"].str.zfill(6)
        log(f"  → {len(df):,}개 종목", 1)
        return df

    # 파일 없으면 API 직접 호출
    log("KRX 종목 파일 없음 → API 수집 중...")
    from datetime import timedelta
    from krx_stock_collector import fetch_all_markets, standardize, recent_business_day
    bas_dd = recent_business_day(offset=-1)
    df_raw = fetch_all_markets(bas_dd)
    if df_raw is None:
        print("❌ KRX 종목 수집 실패. krx_stock_collector.py를 먼저 실행하세요.")
        sys.exit(1)
    df = standardize(df_raw)[["종목코드", "종목명", "시장구분"]].dropna(subset=["종목코드"])
    log(f"  → {len(df):,}개 종목", 1)
    return df


# ═══════════════════════════════════════════════════════════════
#  STEP 2: DART 기업코드 매핑
# ═══════════════════════════════════════════════════════════════

def get_dart_corp_map() -> dict:
    """
    DART corpCode.xml (zip)을 다운로드해서 {stock_code: corp_code} 딕셔너리 반환.
    로컬 캐시(dart_corpcode.csv)가 있으면 재사용.
    """
    cache = "dart_corpcode.csv"
    if os.path.exists(cache):
        df = pd.read_csv(cache, dtype=str)
        log(f"기업코드 캐시 로드: {len(df):,}개")
        return dict(zip(df["stock_code"], df["corp_code"]))

    log("DART 기업코드 다운로드 중...")
    url = f"{DART_BASE}/corpCode.xml?crtfc_key={DART_API_KEY}"
    r = requests.get(url, timeout=30)
    r.raise_for_status()

    with zipfile.ZipFile(io.BytesIO(r.content)) as z:
        xml_name = [n for n in z.namelist() if n.endswith(".xml")][0]
        tree = ET.parse(z.open(xml_name))

    rows = []
    for corp in tree.getroot().iter("list"):
        stock_code = corp.findtext("stock_code", "").strip()
        corp_code  = corp.findtext("corp_code", "").strip()
        corp_name  = corp.findtext("corp_name", "").strip()
        if stock_code:  # 상장사만
            rows.append({"stock_code": stock_code, "corp_code": corp_code,
                         "corp_name": corp_name})

    df = pd.DataFrame(rows)
    df.to_csv(cache, index=False, encoding="utf-8-sig")
    log(f"  → {len(df):,}개 상장 기업코드 저장", 1)
    return dict(zip(df["stock_code"], df["corp_code"]))


# ═══════════════════════════════════════════════════════════════
#  STEP 3: DART API 호출
# ═══════════════════════════════════════════════════════════════

def dart_get(endpoint: str, params: dict) -> list:
    """DART API GET 요청 → OutBlock_1 리스트 반환. 실패 시 빈 리스트."""
    params["crtfc_key"] = DART_API_KEY
    try:
        r = requests.get(f"{DART_BASE}/{endpoint}", params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        status = data.get("status", "")
        if status != "000":
            return []
        return data.get("list", [])
    except Exception as e:
        log(f"    API 오류({endpoint}): {e}", 3)
        return []


def fetch_multi_acnt(corp_codes: list, year: int) -> pd.DataFrame:
    """
    fnlttMultiAcnt: 다중 기업 주요계정
    반환: 매출액, 영업이익, 당기순이익, 자산총계, 부채총계, 자본총계
    """
    rows = []
    for i in range(0, len(corp_codes), BATCH_SIZE):
        batch = corp_codes[i : i + BATCH_SIZE]
        result = dart_get("fnlttMultiAcnt.json", {
            "corp_code":  ",".join(batch),
            "bsns_year":  str(year),
            "reprt_code": REPRT_CODE,
        })
        rows.extend(result)
        time.sleep(REQUEST_DELAY)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    # 계정명 표준화
    df["acnt_std"] = df.get("account_nm", pd.Series(dtype=str)).map(
        lambda x: ACNT_MAP.get(str(x).strip(), str(x).strip())
    )
    # 피벗: corp_code × 계정명
    df_pivot = (
        df[df["acnt_std"].isin(ACNT_MAP.values())]
        .pivot_table(
            index="corp_code",
            columns="acnt_std",
            values="thstrm_amount",   # 당기금액
            aggfunc="first",
        )
        .reset_index()
    )
    df_pivot["bsns_year"] = str(year)
    return df_pivot


def fetch_cmpny_indx(corp_codes: list, year: int) -> pd.DataFrame:
    """
    fnlttCmpnyIndx: 다중 기업 재무비율 (수익성/안정성/성장성/활동성)
    반환: idx_nm별로 피벗된 DataFrame
    """
    all_rows = []
    for cl_code in IDX_CL_CODES:
        for i in range(0, len(corp_codes), BATCH_SIZE):
            batch = corp_codes[i : i + BATCH_SIZE]
            result = dart_get("fnlttCmpnyIndx.json", {
                "corp_code":   ",".join(batch),
                "bsns_year":   str(year),
                "reprt_code":  REPRT_CODE,
                "idx_cl_code": cl_code,
            })
            all_rows.extend(result)
            time.sleep(REQUEST_DELAY)

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    df["idx_val"] = pd.to_numeric(df.get("idx_val", pd.Series(dtype=str)), errors="coerce")

    df_pivot = (
        df[df["idx_val"].notna()]
        .pivot_table(
            index="corp_code",
            columns="idx_nm",
            values="idx_val",
            aggfunc="first",
        )
        .reset_index()
    )
    df_pivot["bsns_year"] = str(year)
    return df_pivot


# ═══════════════════════════════════════════════════════════════
#  STEP 4: 연도별 수집 루프
# ═══════════════════════════════════════════════════════════════

def collect_year(corp_codes: list, year: int) -> tuple[pd.DataFrame, pd.DataFrame]:
    """단일 연도 재무데이터 수집 (체크포인트 재활용)."""

    # ── 주요계정 ──
    df_acnt = load_checkpoint(year, "acnt")
    if df_acnt is not None:
        log(f"  [{year}] 주요계정 체크포인트 로드 ({len(df_acnt)}건)", 1)
    else:
        log(f"  [{year}] 주요계정 수집 중...", 1)
        df_acnt = fetch_multi_acnt(corp_codes, year)
        if not df_acnt.empty:
            save_checkpoint(df_acnt, year, "acnt")
            log(f"    → {len(df_acnt)}개 기업", 2)

    # ── 재무비율 ──
    df_indx = load_checkpoint(year, "indx")
    if df_indx is not None:
        log(f"  [{year}] 재무비율 체크포인트 로드 ({len(df_indx)}건)", 1)
    else:
        log(f"  [{year}] 재무비율 수집 중...", 1)
        df_indx = fetch_cmpny_indx(corp_codes, year)
        if not df_indx.empty:
            save_checkpoint(df_indx, year, "indx")
            log(f"    → {len(df_indx)}개 기업", 2)

    return df_acnt, df_indx


# ═══════════════════════════════════════════════════════════════
#  STEP 5: Excel 저장
# ═══════════════════════════════════════════════════════════════

HEADER_COLORS = {
    "재무제표": "1F4E79",
    "재무비율": "145A32",
    "기업정보": "6E2F77",
}


def style_ws(ws, df, color_key="재무제표"):
    color    = HEADER_COLORS.get(color_key, "1F4E79")
    hdr_fill = PatternFill("solid", fgColor=color)
    hdr_font = Font(bold=True, color="FFFFFF", size=9, name="맑은 고딕")
    alt_fill = PatternFill("solid", fgColor="EEF3FA")
    thin     = Side(style="thin", color="D0D0D0")
    bdr      = Border(left=thin, right=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

        max_w = sum(2 if ord(c) > 127 else 1 for c in str(col)) + 2
        for v in df[col].astype(str):
            w = sum(2 if ord(c) > 127 else 1 for c in v)
            if w > max_w:
                max_w = w
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 1, 30)

    ws.row_dimensions[1].height = 22

    for ri in range(2, min(len(df) + 2, 1048576)):
        for ci in range(1, len(df.columns) + 1):
            cell        = ws.cell(row=ri, column=ci)
            cell.font   = Font(size=8, name="맑은 고딕")
            cell.border = bdr
            if ri % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions


def save_excel(df_info: pd.DataFrame,
               df_fin: pd.DataFrame,
               df_ratio: pd.DataFrame,
               path: str):
    """
    시트 구성 (모두 행별 형태: 기업×연도 = 1행)
      재무제표 : 종목코드 | 종목명 | 시장구분 | 연도 | 매출액 | 영업이익 | …
      재무비율 : 종목코드 | 종목명 | 시장구분 | 연도 | ROE | 부채비율 | …
      기업정보 : 종목코드 | 종목명 | 시장구분 | DART기업코드
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # ── 시트1: 재무제표 (행별) ──
        if not df_fin.empty:
            df_fin.to_excel(writer, sheet_name="재무제표", index=False)
            style_ws(writer.sheets["재무제표"], df_fin, "재무제표")
            log(f"  ✓ [재무제표] {len(df_fin):,}행 ({df_fin['종목코드'].nunique():,}개 기업 × {df_fin['연도'].nunique()}개 연도)", 1)

        # ── 시트2: 재무비율 (행별) ──
        if not df_ratio.empty:
            df_ratio.to_excel(writer, sheet_name="재무비율", index=False)
            style_ws(writer.sheets["재무비율"], df_ratio, "재무비율")
            log(f"  ✓ [재무비율] {len(df_ratio):,}행", 1)

        # ── 시트3: 기업정보 ──
        df_info.to_excel(writer, sheet_name="기업정보", index=False)
        style_ws(writer.sheets["기업정보"], df_info, "기업정보")
        log(f"  ✓ [기업정보] {len(df_info):,}개 기업", 1)


# ═══════════════════════════════════════════════════════════════
#  메인
# ═══════════════════════════════════════════════════════════════

def main():
    print()
    print("=" * 65)
    print("  📊 DART 재무제표 수집기 v1.0")
    print(f"  수집기간  : {TARGET_YEARS[0]}~{TARGET_YEARS[-1]}년 ({len(TARGET_YEARS)}년)")
    print(f"  보고서구분 : 사업보고서 (reprt_code={REPRT_CODE})")
    print(f"  저장파일  : {OUTPUT_FILE}")
    print("=" * 65)

    # ── 1. KRX 종목 로드 ──────────────────────────────────────
    print()
    log("▶ STEP 1: KRX 종목 로드")
    df_krx = load_krx_stocks()

    # ── 2. DART 기업코드 매핑 ─────────────────────────────────
    print()
    log("▶ STEP 2: DART 기업코드 매핑")
    corp_map = get_dart_corp_map()    # {stock_code: corp_code}

    # 매핑 적용
    df_krx["corp_code"] = df_krx["종목코드"].map(corp_map)
    df_krx = df_krx.dropna(subset=["corp_code"])
    corp_codes = df_krx["corp_code"].tolist()
    log(f"  → {len(corp_codes):,}개 기업 DART 매핑 완료", 1)

    if not corp_codes:
        print("❌ 매핑된 기업코드가 없습니다.")
        sys.exit(1)

    # ── 3. 연도별 수집 ────────────────────────────────────────
    print()
    log(f"▶ STEP 3: 재무데이터 수집 ({len(TARGET_YEARS)}개 연도 × {len(corp_codes):,}개 기업)")
    log(f"  예상 API 호출 수: ~{len(TARGET_YEARS) * (len(corp_codes) // BATCH_SIZE + 1) * 5:,}회", 1)
    log(f"  체크포인트 폴더: {CHECKPOINT_DIR}/", 1)

    acnt_frames  = []
    ratio_frames = []

    for year in TARGET_YEARS:
        df_a, df_r = collect_year(corp_codes, year)
        if not df_a.empty:
            acnt_frames.append(df_a)
        if not df_r.empty:
            ratio_frames.append(df_r)

    # ── 4. Long 형태 정리 (기업 × 연도 = 1행) ────────────────
    print()
    log("▶ STEP 4: 데이터 정리 중...")

    df_info = df_krx[["종목코드", "종목명", "시장구분", "corp_code"]].copy()

    def build_long(frames: list) -> pd.DataFrame:
        """연도별 수집 결과를 단순 합치기 (기업×연도 1행)"""
        if not frames:
            return pd.DataFrame()
        df = pd.concat(frames, ignore_index=True)
        df = df.rename(columns={"bsns_year": "연도"})
        return df

    df_fin_long   = build_long(acnt_frames)
    df_ratio_long = build_long(ratio_frames)

    # 기업정보 병합
    if not df_fin_long.empty:
        df_fin_wide   = df_info.merge(df_fin_long,   on="corp_code", how="left") \
                               .sort_values(["종목코드", "연도"]).reset_index(drop=True)
    else:
        df_fin_wide = pd.DataFrame()

    if not df_ratio_long.empty:
        df_ratio_wide = df_info.merge(df_ratio_long, on="corp_code", how="left") \
                               .sort_values(["종목코드", "연도"]).reset_index(drop=True)
    else:
        df_ratio_wide = pd.DataFrame()

    # ── 5. 저장 ───────────────────────────────────────────────
    print()
    log("▶ STEP 5: 엑셀 저장 중...")
    save_excel(df_info, df_fin_wide, df_ratio_wide, OUTPUT_FILE)

    print()
    print("=" * 65)
    print(f"  ✅ 완료!  →  {OUTPUT_FILE}")
    print(f"  기업 수   : {len(df_info):,}개")
    print(f"  수집기간  : {TARGET_YEARS[0]}~{TARGET_YEARS[-1]}년")
    print("=" * 65)
    print()


if __name__ == "__main__":
    missing = [p for p in ("requests", "pandas", "openpyxl")
               if not __import__("importlib").util.find_spec(p)]
    if missing:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing)
        print("설치 완료. 다시 실행해 주세요.\n")
        sys.exit(0)

    main()
