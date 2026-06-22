"""
KRX 전체 종목 수집기 v2.0
────────────────────────────────────────────────────────────────
KOSPI / KOSDAQ / KONEX 전체 종목명·코드·시세를 수집하여 엑셀로 저장

API: https://data-dbg.krx.co.kr/svc/apis/sto/{endpoint}?basDd=YYYYMMDD
     Header → AUTH_KEY: {key}

응답 구조:
  {"OutBlock_1": [
      {"BAS_DD":"...","ISU_CD":"...","ISU_NM":"...","MKT_NM":"...",
       "SECT_TP_NM":"...","TDD_CLSPRC":"...","CMPPREVDD_PRC":"...",
       "FLUC_RT":"...","TDD_OPNPRC":"...","TDD_HGPRC":"...","TDD_LWPRC":"...",
       "ACC_TRDVOL":"...","ACC_TRDVAL":"...","MKTCAP":"...","LIST_SHRS":"..."},
      ...
  ]}

실행 방법:
    pip install requests pandas openpyxl
    python krx_stock_collector.py
────────────────────────────────────────────────────────────────
"""

import sys
import requests
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
#  설정
# ═══════════════════════════════════════════════════════════════

KRX_OPEN_API_KEY = "9D33DFBE2DF54DB99E14C04650F78BE8BB86E937"
OUTPUT_FILE = f"KRX_종목목록_{datetime.today().strftime('%Y%m%d')}.xlsx"

# ── KRX Open API ────────────────────────────────────────────────
API_BASE = "https://data-dbg.krx.co.kr"

# 시장별 엔드포인트
MARKET_ENDPOINTS = {
    "KOSPI":  "/svc/apis/sto/stk_bydd_trd",    # 유가증권 일별시세
    "KOSDAQ": "/svc/apis/sto/ksq_bydd_trd",    # 코스닥 일별시세
}

# ── KRX 정보데이터시스템 폴백 ────────────────────────────────────
KRX_GEN_URL  = "http://data.krx.co.kr/comm/fileDn/GenerateOTP/generate.cmd"
KRX_DOWN_URL = "http://data.krx.co.kr/comm/fileDn/download_csv/download.cmd"
KRX_BROWSER_HEADERS = {
    "Referer": "http://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd?menuId=MDC0201020101",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
}

# ═══════════════════════════════════════════════════════════════
#  API 응답 필드 → 표준 컬럼명 매핑
# ═══════════════════════════════════════════════════════════════

FIELD_MAP = {
    "BAS_DD":        "기준일",
    "ISU_CD":        "종목코드",
    "ISU_NM":        "종목명",
    "MKT_NM":        "시장구분",
    "SECT_TP_NM":    "소속부",
    "TDD_CLSPRC":    "종가",
    "CMPPREVDD_PRC": "전일대비",
    "FLUC_RT":       "등락률",
    "TDD_OPNPRC":    "시가",
    "TDD_HGPRC":     "고가",
    "TDD_LWPRC":     "저가",
    "ACC_TRDVOL":    "거래량",
    "ACC_TRDVAL":    "거래대금",
    "MKTCAP":        "시가총액",
    "LIST_SHRS":     "상장주식수",
}

# 표시 순서
DISPLAY_ORDER = [
    "기준일", "종목코드", "종목명", "시장구분", "소속부",
    "종가", "전일대비", "등락률",
    "시가", "고가", "저가",
    "거래량", "거래대금", "시가총액", "상장주식수",
]

# 숫자형으로 변환할 컬럼
NUM_COLS = {"종가", "전일대비", "등락률", "시가", "고가", "저가",
            "거래량", "거래대금", "시가총액", "상장주식수"}

# ═══════════════════════════════════════════════════════════════
#  유틸리티
# ═══════════════════════════════════════════════════════════════

def log(msg: str, indent: int = 0):
    print("  " * indent + msg)


def recent_business_day(offset: int = 0) -> str:
    """
    최근 영업일 반환 (주말 건너뜀).
    offset=0: 오늘 기준 가장 최근 영업일
    offset=-1: 그 전 영업일, offset=-2: 그보다 하루 전 ...
    """
    d = datetime.today()
    while d.weekday() >= 5:          # 오늘이 주말이면 금요일로
        d -= timedelta(days=1)
    for _ in range(abs(offset)):     # 추가 오프셋
        d -= timedelta(days=1)
        while d.weekday() >= 5:
            d -= timedelta(days=1)
    return d.strftime("%Y%m%d")


# ═══════════════════════════════════════════════════════════════
#  방법 1: KRX Open API
# ═══════════════════════════════════════════════════════════════

def fetch_market(market_name: str, endpoint: str, bas_dd: str) -> pd.DataFrame | None:
    """
    단일 시장 종목 수집.
    GET {API_BASE}{endpoint}?basDd={bas_dd}
    Header: AUTH_KEY
    응답: {"OutBlock_1": [...]}
    """
    url = API_BASE + endpoint
    headers = {
        "AUTH_KEY": KRX_OPEN_API_KEY,
        "Accept":   "application/json",
    }

    try:
        r = requests.get(url, params={"basDd": bas_dd}, headers=headers, timeout=20)

        if r.status_code in (403, 404):
            log(f"  [{market_name}] {r.status_code} — 엔드포인트 없음 또는 권한 없음", 1)
            return None

        r.raise_for_status()

        if not r.content.strip():
            log(f"  [{market_name}] 빈 응답", 1)
            return None

        data = r.json()

        # OutBlock_1 추출
        rows = data.get("OutBlock_1")
        if not rows or not isinstance(rows, list):
            log(f"  [{market_name}] OutBlock_1 없음 또는 비어있음 (키: {list(data.keys())})", 1)
            return None

        df = pd.DataFrame(rows)
        df["_market"] = market_name     # 시장명 보정용 임시 컬럼
        log(f"  [{market_name}] ✓ {len(df):,}개 종목", 1)
        return df

    except Exception as e:
        log(f"  [{market_name}] 오류: {e}", 1)
        return None


def fetch_all_markets(bas_dd: str) -> pd.DataFrame | None:
    """전체 시장 수집 후 합치기. 데이터 없으면 None."""
    frames = []
    for market_name, endpoint in MARKET_ENDPOINTS.items():
        df = fetch_market(market_name, endpoint, bas_dd)
        if df is not None and not df.empty:
            frames.append(df)

    if not frames:
        return None

    return pd.concat(frames, ignore_index=True)


# ═══════════════════════════════════════════════════════════════
#  방법 2: data.krx.co.kr OTP → CSV 폴백
# ═══════════════════════════════════════════════════════════════

def fetch_via_dataportal() -> pd.DataFrame | None:
    """KRX 정보데이터시스템 OTP 방식으로 전종목 수집."""
    log("▶ KRX 정보데이터시스템(폴백) 시도 중...")
    try:
        r = requests.get(
            KRX_GEN_URL,
            params={"mktId": "ALL", "share": "1", "csvxls_isNo": "false",
                    "name": "fileDown", "url": "dbms/MDC/STAT/standard/MDCSTAT01901"},
            headers=KRX_BROWSER_HEADERS,
            timeout=20,
        )
        r.raise_for_status()

        r2 = requests.post(
            KRX_DOWN_URL,
            data={"code": r.content},
            headers=KRX_BROWSER_HEADERS,
            timeout=30,
        )
        r2.raise_for_status()

        df = pd.read_csv(BytesIO(r2.content), encoding="cp949")
        if df.empty:
            log("  빈 CSV 반환", 1)
            return None

        # 포털 CSV 컬럼 → 표준 컬럼명 변환
        portal_map = {
            "단축코드": "종목코드", "한글 종목약명": "종목명", "한글종목명": "종목명",
            "시장구분": "시장구분", "소속부": "소속부",
            "액면가": "액면가", "상장주식수": "상장주식수",
            "종가": "종가", "대비": "전일대비", "등락률": "등락률",
            "거래량": "거래량", "거래대금": "거래대금", "시가총액": "시가총액",
            "상장일": "상장일",
        }
        df = df.rename(columns={k: v for k, v in portal_map.items() if k in df.columns})
        df["기준일"] = datetime.today().strftime("%Y%m%d")
        log(f"  ✓ {len(df):,}개 종목 수집 완료", 1)
        return df

    except Exception as e:
        log(f"  폴백 실패: {e}", 1)
        return None


# ═══════════════════════════════════════════════════════════════
#  데이터 정제
# ═══════════════════════════════════════════════════════════════

def standardize(df: pd.DataFrame) -> pd.DataFrame:
    """컬럼명 표준화, 숫자 변환, 컬럼 순서 정렬."""

    # 1) 컬럼명 매핑 (API 필드명 → 한글)
    df = df.rename(columns={k: v for k, v in FIELD_MAP.items() if k in df.columns})

    # 2) 시장구분 보정: MKT_NM이 없으면 _market 컬럼 사용
    if "시장구분" not in df.columns and "_market" in df.columns:
        df["시장구분"] = df["_market"]
    df = df.drop(columns=["_market"], errors="ignore")

    # 3) 숫자형 변환 (콤마·하이픈 제거)
    for col in NUM_COLS:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                    .str.replace(",", "", regex=False)
                    .str.strip()
                    .replace({"-": None, "--": None, "": None})
            )
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # 4) 종목코드 6자리 패딩
    if "종목코드" in df.columns:
        df["종목코드"] = df["종목코드"].astype(str).str.strip().str.zfill(6)

    # 5) 컬럼 순서
    ordered = [c for c in DISPLAY_ORDER if c in df.columns]
    rest    = [c for c in df.columns if c not in ordered]
    return df[ordered + rest]


def split_by_market(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """시장구분 컬럼 기준으로 딕셔너리 분리."""
    if "시장구분" not in df.columns:
        return {}

    keywords = {
        "KOSPI":  ["KOSPI", "유가증권"],
        "KOSDAQ": ["KOSDAQ", "코스닥"],
        "KONEX":  ["KONEX", "코넥스"],
    }
    result = {}
    matched_idx = pd.Index([])

    for name, kws in keywords.items():
        mask = df["시장구분"].astype(str).str.contains("|".join(kws), case=False, na=False)
        sub = df[mask].copy()
        if not sub.empty:
            result[name] = sub
            matched_idx = matched_idx.union(sub.index)

    other = df[~df.index.isin(matched_idx)]
    if not other.empty:
        result["기타"] = other

    return result


# ═══════════════════════════════════════════════════════════════
#  엑셀 저장 & 스타일
# ═══════════════════════════════════════════════════════════════

HEADER_COLOR = {
    "전체":   "1F4E79",
    "KOSPI":  "1F4E79",
    "KOSDAQ": "145A32",
    "KONEX":  "6E2F77",
    "기타":   "4A4A4A",
}


def style_sheet(ws, df: pd.DataFrame, sheet_name: str):
    color     = HEADER_COLOR.get(sheet_name, "1F4E79")
    hdr_fill  = PatternFill("solid", fgColor=color)
    hdr_font  = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
    alt_fill  = PatternFill("solid", fgColor="EEF3FA")
    thin      = Side(style="thin", color="D0D0D0")
    bdr       = Border(left=thin, right=thin, bottom=thin)

    num_idx = {i + 1 for i, c in enumerate(df.columns) if c in NUM_COLS}

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 열 너비: 한글 2글자 환산
        max_w = sum(2 if ord(c) > 127 else 1 for c in str(col)) + 2
        for v in df[col].astype(str):
            w = sum(2 if ord(c) > 127 else 1 for c in v)
            if w > max_w:
                max_w = w
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 1, 40)

    ws.row_dimensions[1].height = 22

    for ri in range(2, len(df) + 2):
        for ci in range(1, len(df.columns) + 1):
            cell      = ws.cell(row=ri, column=ci)
            cell.font = Font(size=9, name="맑은 고딕")
            cell.border = bdr
            if ri % 2 == 0:
                cell.fill = alt_fill
            if ci in num_idx:
                cell.alignment  = Alignment(horizontal="right")
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_excel(df_all: pd.DataFrame, market_dfs: dict[str, pd.DataFrame], path: str):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # 전체 시트
        df_all.to_excel(writer, sheet_name="전체", index=False)
        style_sheet(writer.sheets["전체"], df_all, "전체")
        log(f"  ✓ [전체] {len(df_all):,}개", 1)

        # 시장별 시트
        for name, df_m in market_dfs.items():
            df_m.to_excel(writer, sheet_name=name, index=False)
            style_sheet(writer.sheets[name], df_m, name)
            log(f"  ✓ [{name}] {len(df_m):,}개", 1)

        # 요약 시트
        wb  = writer.book
        ws2 = wb.create_sheet("📊 요약")
        rows = [
            ["KRX 종목 수집 요약", ""],
            ["수집일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["기준일",  df_all["기준일"].iloc[0] if "기준일" in df_all.columns else "-"],
            ["", ""],
            ["시장", "종목수"],
        ]
        for n, d in market_dfs.items():
            rows.append([n, len(d)])
        rows.append(["합계", len(df_all)])

        for ri, row in enumerate(rows, 1):
            for ci, val in enumerate(row, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                if ri == 1:
                    cell.font = Font(bold=True, size=14, color="1F4E79", name="맑은 고딕")
                elif ri == 5:
                    cell.fill      = PatternFill("solid", fgColor="1F4E79")
                    cell.font      = Font(bold=True, color="FFFFFF", name="맑은 고딕")
                    cell.alignment = Alignment(horizontal="center")
                elif row[0] == "합계":
                    cell.font = Font(bold=True, name="맑은 고딕")
                else:
                    cell.font = Font(size=10, name="맑은 고딕")

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 14


# ═══════════════════════════════════════════════════════════════
#  메인
# ═══════════════════════════════════════════════════════════════

def main():
    bas_dd = recent_business_day(offset=-1)   # 항상 전일 영업일 기준

    print()
    print("=" * 60)
    print("  📈 KRX 전체 종목 수집기 v2.0")
    print(f"  기준일   : {bas_dd}")
    print(f"  저장파일 : {OUTPUT_FILE}")
    print("=" * 60)

    df_raw = None
    source = ""

    # ── 1순위: KRX Open API (data-dbg.krx.co.kr) ──────────────
    print()
    log(f"▶ KRX Open API 수집 중... (basDd={bas_dd})")
    df_raw = fetch_all_markets(bas_dd)

    if df_raw is None or df_raw.empty:
        # 이틀 전 영업일로 재시도 (공휴일 등)
        prev_dd = recent_business_day(offset=-2)
        log(f"\n  → 데이터 없음. 이틀 전({prev_dd})로 재시도...")
        df_raw = fetch_all_markets(prev_dd)
        if df_raw is not None and not df_raw.empty:
            bas_dd = prev_dd

    if df_raw is not None and not df_raw.empty:
        source = "openapi"

    # ── 2순위: data.krx.co.kr OTP 폴백 ────────────────────────
    if not source:
        print()
        df_raw = fetch_via_dataportal()
        if df_raw is not None and not df_raw.empty:
            source = "portal"

    # ── 수집 실패 ──────────────────────────────────────────────
    if not source or df_raw is None or df_raw.empty:
        print()
        print("❌ 종목 데이터를 수집하지 못했습니다.")
        print()
        print("   수동 확인 (터미널에서 실행):")
        print(f'   curl -H "AUTH_KEY: {KRX_OPEN_API_KEY}" \\')
        print(f'        "{API_BASE}/svc/apis/sto/stk_bydd_trd?basDd={bas_dd}"')
        sys.exit(1)

    # ── 정제 & 분리 ────────────────────────────────────────────
    print()
    log("▶ 데이터 정제 중...")
    df_std     = standardize(df_raw)
    market_dfs = split_by_market(df_std)

    # ── 엑셀 저장 ──────────────────────────────────────────────
    print()
    log("▶ 엑셀 저장 중...")
    save_excel(df_std, market_dfs, OUTPUT_FILE)

    print()
    print("=" * 60)
    print(f"  ✅ 완료!  →  {OUTPUT_FILE}")
    print(f"  총 {len(df_std):,}개 종목  |  수집출처: {source}")
    print("=" * 60)
    print()


if __name__ == "__main__":
    missing = [p for p in ("requests", "pandas", "openpyxl") if not __import__("importlib").util.find_spec(p)]
    if missing:
        import subprocess
        print(f"필수 패키지 설치 중: {', '.join(missing)}")
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing)
        print("설치 완료. 다시 실행해 주세요.\n")
        sys.exit(0)

    main()
